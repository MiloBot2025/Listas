import os, time, shutil
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from scripts.common.normalizers import normalize_stock_026, map_026_to_25
from scripts.common.downloader import http_download
from scripts.common.hashio import sha256_of_file

def download_stock(cfg, dest_dir: str):
    method = cfg["source"].get("method","http").lower()
    url = cfg["source"]["url"]
    ts = datetime.now(ZoneInfo("America/Argentina/Buenos_Aires")).strftime("%Y%m%d_%H%M%S")
    out = os.path.join(dest_dir, f"RAW_IMSA_{ts}.xlsx")

    if method == "http":
        http_download(url, out)
        return out, sha256_of_file(out)

    # Selenium
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    dl_dir = os.path.abspath(os.path.join(dest_dir, "_dl")); os.makedirs(dl_dir, exist_ok=True)
    options = Options()
    options.add_argument("--headless=new"); options.add_argument("--no-sandbox"); options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("prefs", {
        "download.default_directory": dl_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    })
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(120)
    try:
        driver.get(url)
        pw_env = cfg["source"].get("password_env")
        if pw_env and os.getenv(pw_env):
            pw_sel = cfg["source"].get("password_selector", "input[type='password']")
            sb_sel = cfg["source"].get("submit_selector", "button[type='submit']")
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, pw_sel))).send_keys(os.getenv(pw_env))
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, sb_sel))).click()
        dl_sel = cfg["source"].get("download_selector")
        if dl_sel:
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, dl_sel))).click()
        # esperar .xlsx
        end = time.time() + 90
        last = None
        while time.time() < end:
            files = [os.path.join(dl_dir, f) for f in os.listdir(dl_dir) if f.lower().endswith((".xlsx",".xls"))]
            if files: last = max(files, key=os.path.getmtime); 
            if last and not last.endswith(".crdownload"): break
            time.sleep(1)
        if not last: raise RuntimeError("IMSA: no se detectó archivo")
        shutil.move(last, out)
        return out, sha256_of_file(out)
    finally:
        driver.quit()

def extract_stock(xlsx_path, cfg):
    wb = load_workbook(xlsx_path, data_only=True)

    # auto headers
    for ws in wb.worksheets:
        hdr = _find_header_row(ws)
        if hdr:
            id_idx, st_idx = hdr["id_col_idx"], hdr["stock_col_idx"]
            rows=[]
            for row in ws.iter_rows(min_row=hdr["row"]+1):
                mid = row[id_idx].value; st = row[st_idx].value
                if mid is None: continue
                v0 = normalize_stock_026(st)
                if cfg.get("solo_con_stock", False):
                    if v0 not in (5,6): continue
                    v = 5
                else:
                    v = map_026_to_25(v0)
                    if v is None: continue
                rows.append({"id": str(mid).strip(), "stock": v})
            return _dedup(rows)

    # fallback
    fb = cfg["fallback"]; ws = wb.active
    start = fb["start_row"]; id_col = ord(fb["id_col"]) - 65; st_col = ord(fb["stock_col"]) - 65
    rows=[]
    for row in ws.iter_rows(min_row=start):
        mid = row[id_col].value; st = row[st_col].value
        if mid is None: continue
        v = map_026_to_25(normalize_stock_026(st))
        if v is None: continue
        rows.append({"id": str(mid).strip(), "stock": v})
    return _dedup(rows)

def _find_header_row(ws):
    for r in range(1, 61):
        vals=[]
        for c in ws[r]:
            v = c.value
            vals.append(v.strip().lower() if isinstance(v,str) else (str(v).lower() if v is not None else ""))
        for i, v in enumerate(vals):
            if any(k in v for k in ["modelo","código","codigo","sku","id"]):
                for j, v2 in enumerate(vals):
                    if any(k in v2 for k in ["stock","disp","exist"]):
                        return {"row": r, "id_col_idx": i, "stock_col_idx": j}
    return None

def _dedup(rows):
    seen={}
    for r in rows: seen[r["id"]] = r
    return list(seen.values())
