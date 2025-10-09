import os
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from scripts.common.normalizers import normalize_stock_026, map_026_to_25
from scripts.common.downloader import http_download
from scripts.common.hashio import sha256_of_file

def download_stock(cfg, dest_dir: str):
    url = cfg["source"]["url"]
    ts = datetime.now(ZoneInfo("America/Argentina/Buenos_Aires")).strftime("%Y%m%d_%H%M%S")
    out = os.path.join(dest_dir, f"RAW_ARS_{ts}.xlsx")
    http_download(url, out)
    return out, sha256_of_file(out)

def extract_stock(xlsx_path, cfg):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[cfg.get("sheet","STOCK")]
    start = cfg["start_row"]; id_col = ord(cfg["id_col"]) - 65; st_col = ord(cfg["stock_col"]) - 65
    out=[]
    for row in ws.iter_rows(min_row=start):
        mid = row[id_col].value; st = row[st_col].value
        if mid is None: continue
        v = map_026_to_25(normalize_stock_026(st))
        if v is None: continue
        out.append({"id": str(mid).strip(), "stock": v})
    seen={}; [seen.setdefault(r["id"], r) for r in out]
    return list(seen.values())
